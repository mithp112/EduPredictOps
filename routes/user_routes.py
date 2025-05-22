from flask import Flask, render_template, request, send_file, jsonify, Blueprint
import pandas as pd
import joblib
import numpy as np
from concurrent.futures import ThreadPoolExecutor
from tensorflow.keras.models import load_model, Model
from tensorflow.keras.layers import Flatten, Dense, LSTM
from tensorflow.keras.saving import register_keras_serializable
from openpyxl.utils import get_column_letter
import io, json
from datetime import date, datetime
from sklearn.metrics import mean_absolute_error, mean_absolute_percentage_error, mean_squared_error, r2_score
import psutil, time
from databaseaction import get_all_schools_and_years, get_submit_data, log_performance_multi, log_performance_single, log_access
from model import load_and_predict


user_blueprint = Blueprint("user", __name__, template_folder = '../templates/user')

# Page home là Page1
@user_blueprint.route('/', strict_slashes=False)
def home():
    return render_template('Page1.html')

# Các route chuyển Page
@user_blueprint.route('/page1')
def page1():
    return render_template('Page1.html')

@user_blueprint.route('/page2')
def page2():
    return render_template('Page2.html')


@user_blueprint.route('/page5')
def page5():
    return render_template('Page5.html')

@user_blueprint.route('/page6', methods=['GET', 'POST'])
def page6():
    subjects_json = request.form.get('subjects')
    subjects_json = subjects_json.replace("'", '"')
    # Kiểm tra xem chuỗi JSON có tồn tại không và đổi thành dictionary
    if subjects_json:
        try:
            subjects = json.loads(subjects_json)
            print(subjects)  # In dictionary đã chuyển đổi
            print(subjects['type'])  # Truy cập giá trị của 'type'
        except json.JSONDecodeError as e:
            print(f"Lỗi JSONDecodeError: {e}")
            subjects = {}
    else:
        print("Không có dữ liệu subjects_json.")
        subjects = {}

    return render_template('Page6.html', subjects=subjects)


def predModel1(ypred): # hiển thị kết quả cho model 1
    pred = pd.DataFrame({
        'Toán_1': pd.Series(ypred[:,0]),
        'Văn_1': pd.Series(ypred[:,1]),
        'Lý_1': pd.Series(ypred[:,2]),
        'Hóa_1': pd.Series(ypred[:,3]),
        'Sinh_1': pd.Series(ypred[:,4]),
        'Sử_1': pd.Series(ypred[:,5]),
        'Địa_1': pd.Series(ypred[:,6]),
        'Anh_1': pd.Series(ypred[:,7]),
        'GDCD_1': pd.Series(ypred[:,8]),
        'Toán_2': pd.Series(ypred[:,9]),
        'Văn_2': pd.Series(ypred[:,10]),
        'Lý_2': pd.Series(ypred[:,11]),
        'Hóa_2': pd.Series(ypred[:,12]),
        'Sinh_2': pd.Series(ypred[:,13]),
        'Sử_2': pd.Series(ypred[:,14]),
        'Địa_2': pd.Series(ypred[:,15]),
        'Anh_2': pd.Series(ypred[:,16]),
        'GDCD_2': pd.Series(ypred[:,17])
    })
    pred = pred.astype('float64')
    return pred

def predModel2(ypred): # hiển thị kết quả cho model 2
    pred = pd.DataFrame({
        'Toán_1': pd.Series(ypred[:,0]),
        'Văn_1': pd.Series(ypred[:,1]),
        'Lý_1': pd.Series(ypred[:,2]),
        'Hóa_1': pd.Series(ypred[:,3]),
        'Sinh_1': pd.Series(ypred[:,4]),
        'Anh_1': pd.Series(ypred[:,5]),
    })
    pred = pred.astype('float64')
    return pred


    
@user_blueprint.route("predict1", methods = ['GET','POST'])
def predict1():
    school_name = request.form.get('schoolName')
    log_access('/predict1', school_name)
    start_time = time.time()  # Bắt đầu đo thời gian
    cpu_start = psutil.cpu_percent(interval=None)  # Lấy CPU ban đầu
    memory_start = psutil.Process().memory_info().rss / (1024 ** 2)
    pred_df1 = None
    pred_df2 = None
    pred_df3 = None
    if request.method == "POST":
        try:
            # Lấy dữ liệu điểm do người dùng nhập vào 
            Toan_1_10 = float(request.form['Toan_1_10'])
            Toan_2_10 = float(request.form['Toan_2_10'])
            Toan_1_11 = float(request.form['Toan_1_11'])
            Toan_2_11 = float(request.form['Toan_2_11'])

            Van_1_10 = float(request.form['Van_1_10'])
            Van_2_10 = float(request.form['Van_2_10'])
            Van_1_11 = float(request.form['Van_1_11'])
            Van_2_11 = float(request.form['Van_2_11'])

            Ly_1_10 = float(request.form['Ly_1_10'])
            Ly_2_10 = float(request.form['Ly_2_10'])
            Ly_1_11 = float(request.form['Ly_1_11'])
            Ly_2_11 = float(request.form['Ly_2_11'])

            Anh_1_10 = float(request.form['Anh_1_10'])
            Anh_2_10 = float(request.form['Anh_2_10'])
            Anh_1_11 = float(request.form['Anh_1_11'])
            Anh_2_11 = float(request.form['Anh_2_11'])

            Su_1_10 = float(request.form['Su_1_10'])
            Su_2_10 = float(request.form['Su_2_10'])
            Su_1_11 = float(request.form['Su_1_11'])
            Su_2_11 = float(request.form['Su_2_11'])

            Dia_1_10 = float(request.form['Dia_1_10'])
            Dia_2_10 = float(request.form['Dia_2_10'])
            Dia_1_11 = float(request.form['Dia_1_11'])
            Dia_2_11 = float(request.form['Dia_2_11'])

            Sinh_1_10 = float(request.form['Sinh_1_10'])
            Sinh_2_10 = float(request.form['Sinh_2_10'])
            Sinh_1_11 = float(request.form['Sinh_1_11'])
            Sinh_2_11 = float(request.form['Sinh_2_11'])

            Hoa_1_10 = float(request.form['Hoa_1_10'])
            Hoa_2_10 = float(request.form['Hoa_2_10'])
            Hoa_1_11 = float(request.form['Hoa_1_11'])
            Hoa_2_11 = float(request.form['Hoa_2_11'])

            GDCD_1_10 = float(request.form['GDCD_1_10'])
            GDCD_2_10 = float(request.form['GDCD_2_10'])
            GDCD_1_11 = float(request.form['GDCD_1_11'])
            GDCD_2_11 = float(request.form['GDCD_2_11'])
            # Lấy dư liệu khó khăn và mồ côi từ checkbox
            orphan = 1 if 'orphan' in request.form else 0
            kios = 2 if 'kios' in request.form else 0
            orphan_and_kios = kios + orphan

            pred_arg = [Toan_1_10,Van_1_10,Ly_1_10,Hoa_1_10,Sinh_1_10,Su_1_10,Dia_1_10,Anh_1_10, GDCD_1_10
                        ,Toan_2_10,Van_2_10,Ly_2_10,Hoa_2_10,Sinh_2_10,Su_2_10,Dia_2_10,Anh_2_10, GDCD_2_10
                        ,Toan_1_11,Van_1_11,Ly_1_11,Hoa_1_11,Sinh_1_11,Su_1_11,Dia_1_11,Anh_1_11, GDCD_1_11
                        ,Toan_2_11,Van_2_11,Ly_2_11,Hoa_2_11,Sinh_2_11,Su_2_11,Dia_2_11,Anh_2_11, GDCD_2_11
                        ,orphan_and_kios]
            
            pred_arg_arr = np.array(pred_arg)
            pred_arg_arr = pred_arg_arr.reshape(1,-1)
            # Model LR được lưu bằng joblib, 2 model còn lại MLP và LSTM được lưu băng Karas
            # Thực hiện dự đoán song song trên cả 3 model
            with ThreadPoolExecutor() as executor:
                future1 = executor.submit(load_and_predict, school_name, "LR_10_11_12.pkl", model_type='joblib',pred_arg_arr=pred_arg_arr)
                future2 = executor.submit(load_and_predict, school_name, "MLP_10_11_12.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                future3 = executor.submit(load_and_predict, school_name,"LSTM_10_11_12.keras", model_type='keras',pred_arg_arr=pred_arg_arr)

                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel1(prediction1)
                pred_df2 = predModel1(prediction2)
                pred_df3 = predModel1(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
            # Đo hiệu năng sau khi xử lý xong
            elapsed_time = time.time() - start_time
            cpu_end = psutil.cpu_percent(interval=None)
            memory_end = psutil.Process().memory_info().rss / (1024 ** 2)

            # Tính toán hiệu năng
            avg_cpu_usage = (cpu_start + cpu_end) / 2
            avg_memory_usage = (memory_start + memory_end) / 2
            throughput = len(pred_arg_arr) / elapsed_time if elapsed_time > 0 else 0

            performance = {}
            performance['timestamp']=datetime.utcnow()
            performance['latency']=round(elapsed_time * 1000, 2),  # Tính độ trễ (ms)
            performance['throughput']=round(throughput, 2),
            performance['cpu_usage']=round(avg_cpu_usage, 2),
            performance['memory_usage']=round(avg_memory_usage, 2),
            log_performance_multi(school_name = school_name, performance=performance)

                

        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
        
    # Render kết quả dự đoán được về Page hiển thị kết quả
    return render_template('Page3.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
)


@user_blueprint.route("/predict2", methods = ['GET','POST'])
def predict2():
    school_name = request.form.get('schoolName')
    log_access('/predict2', school_name)
    start_time = time.time()  # Bắt đầu đo thời gian
    cpu_start = psutil.cpu_percent(interval=None)  # Lấy CPU ban đầu
    memory_start = psutil.Process().memory_info().rss / (1024 ** 2)
    start_time = time.time()
    if request.method == "POST":
        try:
            Toan_1_10 = float(request.form['Toan_1_10'])
            Toan_2_10 = float(request.form['Toan_2_10'])
            Toan_1_11 = float(request.form['Toan_1_11'])
            Toan_2_11 = float(request.form['Toan_2_11'])
            Toan_1_12 = float(request.form['Toan_1_12'])
            Toan_2_12 = float(request.form['Toan_2_12'])

            Van_1_10 = float(request.form['Van_1_10'])
            Van_2_10 = float(request.form['Van_2_10'])
            Van_1_11 = float(request.form['Van_1_11'])
            Van_2_11 = float(request.form['Van_2_11'])
            Van_1_12 = float(request.form['Van_1_12'])
            Van_2_12 = float(request.form['Van_2_12'])

            Ly_1_10 = float(request.form['Ly_1_10'])
            Ly_2_10 = float(request.form['Ly_2_10'])
            Ly_1_11 = float(request.form['Ly_1_11'])
            Ly_2_11 = float(request.form['Ly_2_11'])
            Ly_1_12 = float(request.form['Ly_1_12'])
            Ly_2_12 = float(request.form['Ly_2_12'])

            Anh_1_10 = float(request.form['Anh_1_10'])
            Anh_2_10 = float(request.form['Anh_2_10'])
            Anh_1_11 = float(request.form['Anh_1_11'])
            Anh_2_11 = float(request.form['Anh_2_11'])
            Anh_1_12 = float(request.form['Anh_1_12'])
            Anh_2_12 = float(request.form['Anh_2_12'])

            Su_1_10 = float(request.form['Su_1_10'])
            Su_2_10 = float(request.form['Su_2_10'])
            Su_1_11 = float(request.form['Su_1_11'])
            Su_2_11 = float(request.form['Su_2_11'])
            Su_1_12 = float(request.form['Su_1_12'])
            Su_2_12 = float(request.form['Su_2_12'])


            Dia_1_10 = float(request.form['Dia_1_10'])
            Dia_2_10 = float(request.form['Dia_2_10'])
            Dia_1_11 = float(request.form['Dia_1_11'])
            Dia_2_11 = float(request.form['Dia_2_11'])
            Dia_1_12 = float(request.form['Dia_1_12'])
            Dia_2_12 = float(request.form['Dia_2_12'])

            Sinh_1_10 = float(request.form['Sinh_1_10'])
            Sinh_2_10 = float(request.form['Sinh_2_10'])
            Sinh_1_11 = float(request.form['Sinh_1_11'])
            Sinh_2_11 = float(request.form['Sinh_2_11'])
            Sinh_1_12 = float(request.form['Sinh_1_12'])
            Sinh_2_12 = float(request.form['Sinh_2_12'])

            Hoa_1_10 = float(request.form['Hoa_1_10'])
            Hoa_2_10 = float(request.form['Hoa_2_10'])
            Hoa_1_11 = float(request.form['Hoa_1_11'])
            Hoa_2_11 = float(request.form['Hoa_2_11'])
            Hoa_1_12 = float(request.form['Hoa_1_12'])
            Hoa_2_12 = float(request.form['Hoa_2_12'])

            GDCD_1_10 = float(request.form['GDCD_1_10'])
            GDCD_2_10 = float(request.form['GDCD_2_10'])
            GDCD_1_11 = float(request.form['GDCD_1_11'])
            GDCD_2_11 = float(request.form['GDCD_2_11'])
            GDCD_1_12 = float(request.form['GDCD_1_12'])
            GDCD_2_12 = float(request.form['GDCD_2_12'])
            orphan = 1 if 'orphan' in request.form else 0
            kios = 2 if 'kios' in request.form else 0
            orphan_and_kios = kios + orphan
            

            pred_arg = [Toan_1_10,Van_1_10,Ly_1_10,Hoa_1_10,Sinh_1_10,Su_1_10,Dia_1_10,Anh_1_10, GDCD_1_10
                        ,Toan_2_10,Van_2_10,Ly_2_10,Hoa_2_10,Sinh_2_10,Su_2_10,Dia_2_10,Anh_2_10, GDCD_2_10
                        ,Toan_1_11,Van_1_11,Ly_1_11,Hoa_1_11,Sinh_1_11,Su_1_11,Dia_1_11,Anh_1_11, GDCD_1_11
                        ,Toan_2_11,Van_2_11,Ly_2_11,Hoa_2_11,Sinh_2_11,Su_2_11,Dia_2_11,Anh_2_11, GDCD_2_11
                        ,Toan_1_12,Van_1_12,Ly_1_12,Hoa_1_12,Sinh_1_12,Su_1_12,Dia_1_12,Anh_1_12, GDCD_1_12
                        ,Toan_2_12,Van_2_12,Ly_2_12,Hoa_2_12,Sinh_2_12,Su_2_12,Dia_2_12,Anh_2_12, GDCD_2_12
                        ,orphan_and_kios]
            
            
            pred_arg_arr = np.array(pred_arg)
            pred_arg_arr = pred_arg_arr.reshape(1,-1)
            nature_checked = True if 'nature' in request.form else False
            subjects = {}

            with ThreadPoolExecutor() as executor:
                if nature_checked:
                    future1 = executor.submit(load_and_predict, school_name, "LR_TN_TN.pkl", model_type='joblib',pred_arg_arr=pred_arg_arr)
                    future2 = executor.submit(load_and_predict, school_name, "MLP_TN_TN.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                    future3 = executor.submit(load_and_predict, school_name, "LSTM_TN_TN.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                    subjects = {'mon1': 'Điểm Lý', 'mon1_1': 'Physics', 'mon2': 'Điểm Hóa', 'mon2_1': 'Chemistry', 'mon3': 'Điểm Sinh', 'mon3_1':  'Biology', 'type':  'TN'}
                else:
                    future1 = executor.submit(load_and_predict, school_name, "LR_TN_XH.pkl", model_type='joblib',pred_arg_arr=pred_arg_arr)
                    future2 = executor.submit(load_and_predict, school_name, "MLP_TN_XH.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                    future3 = executor.submit(load_and_predict, school_name, "LSTM_TN_XH.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                    subjects = {'mon1': 'Điểm Sử', 'mon1_1': 'History', 'mon2': 'Điểm Địa', 'mon2_1': 'Geography', 'mon3': 'Điểm GDCD', 'mon3_1':  'Civic Education', 'type':  'XH'}
                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel2(prediction1)
                pred_df2 = predModel2(prediction2)
                pred_df3 = predModel2(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
            # Đo hiệu năng sau khi xử lý xong
            elapsed_time = time.time() - start_time
            cpu_end = psutil.cpu_percent(interval=None)
            memory_end = psutil.Process().memory_info().rss / (1024 ** 2)

            # Tính toán hiệu năng
            avg_cpu_usage = (cpu_start + cpu_end) / 2
            avg_memory_usage = (memory_start + memory_end) / 2
            throughput = len(pred_arg_arr) / elapsed_time if elapsed_time > 0 else 0

            performance = {}
            performance['timestamp']=datetime.utcnow()
            performance['latency']=round(elapsed_time * 1000, 2),  # Tính độ trễ (ms)
            performance['throughput']=round(throughput, 2),
            performance['cpu_usage']=round(avg_cpu_usage, 2),
            performance['memory_usage']=round(avg_memory_usage, 2),
            log_performance_single(school_name = school_name, performance=performance)
                
        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
    return render_template('Page4.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
    subjects=subjects
)



@user_blueprint.route("/predict_excel1", methods=['GET', 'POST'])
def predict_excel1():
    school_name = request.form.get('schoolName')
    log_access('/predict_excel1', school_name)
    start_time = time.time()  # Bắt đầu đo thời gian
    cpu_start = psutil.cpu_percent(interval=None)  # Lấy CPU ban đầu
    memory_start = psutil.Process().memory_info().rss / (1024 ** 2)
    if request.method == "POST":
        try:
            # Đọc file excel và lấy dữ liệu, bỏ hàng đầu tiên, lấy cột name đâu tiên
            excel_file = request.files['excel_file']
            df = pd.read_excel(excel_file,header=None, skiprows=1) 
            names = df.iloc[:, 0].tolist()
            df = df.drop(columns=0) 
            df = df.iloc[:, :37]
            pred_arg_arr = df.to_numpy()
            def load_and_predict(model_path, model_type):
                if model_type == 'keras':
                    model = load_model(f"{school_name}/Models/{model_path}")  # Load model Keras
                else:
                    with open(f"{school_name}/Models/{model_path}", 'rb') as model_file:
                        model = joblib.load(model_file)  # Load model với joblib

                if model_type == 'keras' and "LSTM" in model_path:  # Kiểm tra xem đây có phải là model LSTM không
                    return model.predict(pred_arg_arr.reshape(pred_arg_arr.shape[0], 1, pred_arg_arr.shape[1]))
                else:
                    return model.predict(pred_arg_arr)

           
            with ThreadPoolExecutor() as executor:
                future1 = executor.submit(load_and_predict, school_name, "LR_10_11_12.pkl", model_type='joblib', pred_arg_arr=pred_arg_arr)
                future2 = executor.submit(load_and_predict, school_name, "MLP_10_11_12.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                future3 = executor.submit(load_and_predict, school_name, "LSTM_10_11_12.keras", model_type='keras',pred_arg_arr=pred_arg_arr)

                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel1(prediction1)
                pred_df2 = predModel1(prediction2)
                pred_df3 = predModel1(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
            # Đo hiệu năng sau khi xử lý xong
            elapsed_time = time.time() - start_time
            cpu_end = psutil.cpu_percent(interval=None)
            memory_end = psutil.Process().memory_info().rss / (1024 ** 2)

            # Tính toán hiệu năng
            avg_cpu_usage = (cpu_start + cpu_end) / 2
            avg_memory_usage = (memory_start + memory_end) / 2
            throughput = len(pred_arg_arr) / elapsed_time if elapsed_time > 0 else 0

            # Lưu hiệu năng vào cơ sở dữ liệu
            performance = {}
            performance['timestamp']=datetime.utcnow()
            performance['latency']=round(elapsed_time * 1000, 2),  # Tính độ trễ (ms)
            performance['throughput']=round(throughput, 2),
            performance['cpu_usage']=round(avg_cpu_usage, 2),
            performance['memory_usage']=round(avg_memory_usage, 2),
            performance['total_predictions']=len(pred_arg_arr)
            log_performance_multi(school_name = school_name, performance=performance)
                
        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
            return {"error": str(e)}

    return render_template('Page3.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
    names=names,
    school_name = school_name
)


@user_blueprint.route('/download_excel1', methods=['POST'])
def download_excel1():
    school_name = request.form.get('schoolName')
    names = request.form.getlist('names')
    kq_Model1 = request.form.getlist('kq_Model1')
    kq_Model2 = request.form.getlist('kq_Model2')
    kq_Model3 = request.form.getlist('kq_Model3')

    names = json.loads(names[0].replace("'", '"'))
    kq_Model1 = json.loads(kq_Model1[0].replace("'", '"'))
    kq_Model2 = json.loads(kq_Model2[0].replace("'", '"'))
    kq_Model3 = json.loads(kq_Model3[0].replace("'", '"'))

    dataLR = {
        "Tên": names,

        # Kết quả từ LR model
        "Điểm Toán I": [kq_Model1['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn I": [kq_Model1['Văn_1'][i] for i in range(len(names))],
        "Điểm Lý I": [kq_Model1['Lý_1'][i] for i in range(len(names))],
        "Điểm Hóa I": [kq_Model1['Hóa_1'][i] for i in range(len(names))],
        "Điểm Sinh I": [kq_Model1['Sinh_1'][i] for i in range(len(names))],
        "Điểm Sử I": [kq_Model1['Sử_1'][i] for i in range(len(names))],
        "Điểm Địa I": [kq_Model1['Địa_1'][i] for i in range(len(names))],
        "Điểm GDCD I": [kq_Model1['GDCD_1'][i] for i in range(len(names))],
        "Điểm Anh I": [kq_Model1['Anh_1'][i] for i in range(len(names))],

        "Điểm Toán II": [kq_Model1['Toán_2'][i] for i in range(len(names))],
        "Điểm Văn II": [kq_Model1['Văn_2'][i] for i in range(len(names))],
        "Điểm Lý II": [kq_Model1['Lý_2'][i] for i in range(len(names))],
        "Điểm Hóa II": [kq_Model1['Hóa_2'][i] for i in range(len(names))],
        "Điểm Sinh II": [kq_Model1['Sinh_2'][i] for i in range(len(names))],
        "Điểm Sử II": [kq_Model1['Sử_2'][i] for i in range(len(names))],
        "Điểm Địa II": [kq_Model1['Địa_2'][i] for i in range(len(names))],
        "Điểm GDCD II": [kq_Model1['GDCD_2'][i] for i in range(len(names))],
        "Điểm Anh II": [kq_Model1['Anh_2'][i] for i in range(len(names))],
    }

    dataMLP = {
        "Tên": names,
        
        # Kết quả từ MLP model
        "Điểm Toán I": [kq_Model2['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn I": [kq_Model2['Văn_1'][i] for i in range(len(names))],
        "Điểm Lý I": [kq_Model2['Lý_1'][i] for i in range(len(names))],
        "Điểm Hóa I": [kq_Model2['Hóa_1'][i] for i in range(len(names))],
        "Điểm Sinh I": [kq_Model2['Sinh_1'][i] for i in range(len(names))],
        "Điểm Sử I": [kq_Model2['Sử_1'][i] for i in range(len(names))],
        "Điểm Địa I": [kq_Model2['Địa_1'][i] for i in range(len(names))],
        "Điểm GDCD I": [kq_Model2['GDCD_1'][i] for i in range(len(names))],
        "Điểm Anh I": [kq_Model2['Anh_1'][i] for i in range(len(names))],

        "Điểm Toán II": [kq_Model2['Toán_2'][i] for i in range(len(names))],
        "Điểm Văn II": [kq_Model2['Văn_2'][i] for i in range(len(names))],
        "Điểm Lý II": [kq_Model2['Lý_2'][i] for i in range(len(names))],
        "Điểm Hóa II": [kq_Model2['Hóa_2'][i] for i in range(len(names))],
        "Điểm Sinh II": [kq_Model2['Sinh_2'][i] for i in range(len(names))],
        "Điểm Sử II": [kq_Model2['Sử_2'][i] for i in range(len(names))],
        "Điểm Địa II": [kq_Model2['Địa_2'][i] for i in range(len(names))],
        "Điểm GDCD II": [kq_Model2['GDCD_2'][i] for i in range(len(names))],
        "Điểm Anh II": [kq_Model2['Anh_2'][i] for i in range(len(names))],
    }

    dataLSTM = {
        "Tên": names,

        # Kết quả từ LSTM model
        "Điểm Toán I": [kq_Model3['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn I": [kq_Model3['Văn_1'][i] for i in range(len(names))],
        "Điểm Lý I": [kq_Model3['Lý_1'][i] for i in range(len(names))],
        "Điểm Hóa I": [kq_Model3['Hóa_1'][i] for i in range(len(names))],
        "Điểm Sinh I": [kq_Model3['Sinh_1'][i] for i in range(len(names))],
        "Điểm Sử I": [kq_Model3['Sử_1'][i] for i in range(len(names))],
        "Điểm Địa I": [kq_Model3['Địa_1'][i] for i in range(len(names))],
        "Điểm GDCD I": [kq_Model3['GDCD_1'][i] for i in range(len(names))],
        "Điểm Anh I": [kq_Model3['Anh_1'][i] for i in range(len(names))],

        "Điểm Toán II": [kq_Model3['Toán_2'][i] for i in range(len(names))],
        "Điểm Văn II": [kq_Model3['Văn_2'][i] for i in range(len(names))],
        "Điểm Lý II": [kq_Model3['Lý_2'][i] for i in range(len(names))],
        "Điểm Hóa II": [kq_Model3['Hóa_2'][i] for i in range(len(names))],
        "Điểm Sinh II": [kq_Model3['Sinh_2'][i] for i in range(len(names))],
        "Điểm Sử II": [kq_Model3['Sử_2'][i] for i in range(len(names))],
        "Điểm Địa II": [kq_Model3['Địa_2'][i] for i in range(len(names))],
        "Điểm GDCD II": [kq_Model3['GDCD_2'][i] for i in range(len(names))],
        "Điểm Anh II": [kq_Model3['Anh_2'][i] for i in range(len(names))],
    }
    
    # Tạo DataFrame
    df1 = pd.DataFrame(dataLR)
    df2 = pd.DataFrame(dataMLP)
    df3 = pd.DataFrame(dataLSTM)

    # Tạo đối tượng Excel trong bộ nhớ
    output = io.BytesIO()


    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='LR', index=False)
        df2.to_excel(writer, sheet_name='MLP', index=False)
        df3.to_excel(writer, sheet_name='LSTM', index=False)
    
    # Tự động điều chỉnh độ rộng của các cột cho từng sheet
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = length + 2  # Thêm khoảng cách để nội dung thoáng hơn

    output.seek(0)

    # Trả về file Excel để tải xuống
    return send_file(output, as_attachment=True, download_name=f'Dự đoán điểm Trung Bình lớp 12 Trường {school_name}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@user_blueprint.route("/predict_excel2", methods=['GET', 'POST'])
def predict_excel2():
    school_name = request.form.get('schoolName')
    log_access('/predict_excel2', school_name)
    start_time = time.time()  # Bắt đầu đo thời gian
    cpu_start = psutil.cpu_percent(interval=None)  # Lấy CPU ban đầu
    memory_start = psutil.Process().memory_info().rss / (1024 ** 2)  # Lấy Memory ban đầu
    if request.method == "POST":
        try:
            excel_file = request.files['excel_file']
            df = pd.read_excel(excel_file,header=None, skiprows=1)  
            names = df.iloc[:, 0].tolist() 
            df = df.drop(columns=0)
            df = df.iloc[:, :55]
            pred_arg_arr = df.to_numpy()


            # Kiểm tra checkbox TN (Tự nhiên) hay XH (Xã hội)
            nature_checked = True if 'nature' in request.form else False
            subjects = {}

            with ThreadPoolExecutor() as executor:
                if nature_checked:
                    future1 = executor.submit(load_and_predict, school_name, "LR_TN_TN.pkl", model_type='joblib',pred_arg_arr=pred_arg_arr)
                    future2 = executor.submit(load_and_predict, school_name, "MLP_TN_TN.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                    future3 = executor.submit(load_and_predict, school_name, "LSTM_TN_TN.keras", model_type='keras',pred_arg_arr=pred_arg_arr)
                    subjects = {'mon1': 'Điểm Lý', 'mon1_1': 'Physics', 'mon2': 'Điểm Hóa', 'mon2_1': 'Chemistry', 'mon3': 'Điểm Sinh', 'mon3_1': 'Biology', 'type': 'TN'}
                else:
                    future1 = executor.submit(load_and_predict, "LR_TN_XH.pkl", model_type='joblib')
                    future2 = executor.submit(load_and_predict, "MLP_TN_XH.keras", model_type='keras')
                    future3 = executor.submit(load_and_predict, "LSTM_TN_XH.keras", model_type='keras')
                    subjects = {'mon1': 'Điểm Sử', 'mon1_1': 'History', 'mon2': 'Điểm Địa', 'mon2_1': 'Geography', 'mon3': 'Điểm GDCD', 'mon3_1': 'Civic Education', 'type': 'XH'}

                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel2(prediction1)
                pred_df2 = predModel2(prediction2)
                pred_df3 = predModel2(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
            # Đo hiệu năng sau khi xử lý xong
            elapsed_time = time.time() - start_time
            cpu_end = psutil.cpu_percent(interval=None)
            memory_end = psutil.Process().memory_info().rss / (1024 ** 2)

            # Tính toán hiệu năng
            avg_cpu_usage = (cpu_start + cpu_end) / 2
            avg_memory_usage = (memory_start + memory_end) / 2
            throughput = len(pred_arg_arr) / elapsed_time if elapsed_time > 0 else 0


            performance = {}
            performance['timestamp']=datetime.utcnow()
            performance['latency']=round(elapsed_time * 1000, 2),  # Tính độ trễ (ms)
            performance['throughput']=round(throughput, 2),
            performance['cpu_usage']=round(avg_cpu_usage, 2),
            performance['memory_usage']=round(avg_memory_usage, 2),
            performance['total_predictions']=len(pred_arg_arr)
            log_performance_multi(school_name = school_name, performance=performance)


        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
            return {"error": str(e)}

    return render_template('Page4.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
    subjects=subjects,
    names=names,
    school_name = school_name
)

@user_blueprint.route('/download_excel2', methods=['POST'])
def download_excel2():
    school_name = request.form.getlist('school_name')
    names = request.form.getlist('names')
    kq_Model1 = request.form.getlist('kq_Model1')
    kq_Model2 = request.form.getlist('kq_Model2')
    kq_Model3 = request.form.getlist('kq_Model3')
    subjects = request.form.getlist('subjects')

    names = json.loads(names[0].replace("'", '"'))
    subjects = json.loads(subjects[0].replace("'", '"'))
    kq_Model1 = json.loads(kq_Model1[0].replace("'", '"'))
    kq_Model2 = json.loads(kq_Model2[0].replace("'", '"'))
    kq_Model3 = json.loads(kq_Model3[0].replace("'", '"'))

    dataLR = {
        "Tên": names,

        # Kết quả từ LR model
        "Điểm Toán": [kq_Model1['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn": [kq_Model1['Văn_1'][i] for i in range(len(names))],
        subjects.get('mon1'): [kq_Model1['Lý_1'][i] for i in range(len(names))],
        subjects.get('mon2'): [kq_Model1['Hóa_1'][i] for i in range(len(names))],
        subjects.get('mon3'): [kq_Model1['Sinh_1'][i] for i in range(len(names))],
        "Điểm Anh": [kq_Model1['Anh_1'][i] for i in range(len(names))],
    }

    dataMLP = {
        "Tên": names,
        
        # Kết quả từ MLP model
        "Điểm Toán": [kq_Model2['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn": [kq_Model2['Văn_1'][i] for i in range(len(names))],
        subjects.get('mon1'): [kq_Model2['Lý_1'][i] for i in range(len(names))],
        subjects.get('mon2'): [kq_Model2['Hóa_1'][i] for i in range(len(names))],
        subjects.get('mon3'): [kq_Model2['Sinh_1'][i] for i in range(len(names))],
        "Điểm Anh": [kq_Model2['Anh_1'][i] for i in range(len(names))],
    }

    dataLSTM = {
        "Tên": names,

        # Kết quả từ LSTM model
        "Điểm Toán": [kq_Model3['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn": [kq_Model3['Văn_1'][i] for i in range(len(names))],
        subjects.get('mon1'): [kq_Model3['Lý_1'][i] for i in range(len(names))],
        subjects.get('mon2'): [kq_Model3['Hóa_1'][i] for i in range(len(names))],
        subjects.get('mon3'): [kq_Model3['Sinh_1'][i] for i in range(len(names))],
        "Điểm Anh": [kq_Model3['Anh_1'][i] for i in range(len(names))]
    }
    
    # Tạo DataFrame
    df1 = pd.DataFrame(dataLR)
    df2 = pd.DataFrame(dataMLP)
    df3 = pd.DataFrame(dataLSTM)

    # Tạo đối tượng Excel trong bộ nhớ
    output = io.BytesIO()


    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='LR', index=False)
        df2.to_excel(writer, sheet_name='MLP', index=False)
        df3.to_excel(writer, sheet_name='LSTM', index=False)
    
    # Tự động điều chỉnh độ rộng của các cột cho từng sheet
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = length + 2  # Thêm khoảng cách để nội dung thoáng hơn

    output.seek(0)

    # Trả về file Excel để tải xuống
    return send_file(output, as_attachment=True, download_name=f'Dự đoán điểm thi Tốt Nghiệp {school_name}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@user_blueprint.route('/get_schools', methods =['GET'])
def get_schools():
    data = get_all_schools_and_years()
    return jsonify(data)

@user_blueprint.route('/get_submit_data', methods =['GET'])
def get_submit_data():
    data = get_submit_data()
    return jsonify(data)